import os
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

# Load environment variables
time_value = int(os.environ["KMVAR_DeepWorkTimeCompleted"])
column_name = os.environ["KMVAR_local_categorySelection"]

# Load the Excel file
file_path = os.path.expanduser("~/Documents/Coding-Projects/Keyboard-Maestro-and-Obsidian/Deep Work Time Tracker.xlsx")
workbook = load_workbook(filename=file_path)
sheet = workbook.active

# Find the first empty row in the specified column
column_index = None
for idx, cell in enumerate(sheet[1], 1):
    if cell.value == column_name:
        column_index = idx
        break

if column_index is None:
    raise ValueError(f"Column '{column_name}' not found in the Excel file.")
else:
    # Identify the date column (assumed to be the one to the left of the specified column)
    date_column_index = column_index - 1

    # Find the first empty row in the specified column
    for row in range(2, sheet.max_row + 2):
        if sheet.cell(row=row, column=column_index).value is None:
            # Insert the time and date
            sheet.cell(row=row, column=date_column_index).value = datetime.now().strftime("%m/%d/%Y %H:%M")
            sheet.cell(row=row, column=column_index).value = time_value
            break

# Iterate through the sheet to find each "Total" column
for col_idx in range(1, sheet.max_column + 1):
    cell_value = sheet.cell(row=1, column=col_idx).value
    if cell_value and "Total" in cell_value:
        # Identify the corresponding date column (assumed to be the one to the left of the value column)
        date_col_idx = col_idx - 2  # Two columns to the left (since it's "Date", then "Value", then "Total")
        if date_col_idx > 0:
            # Target the first row under the header (row 2) for the sum formula
            value_range = f"{sheet.cell(row=2, column=col_idx-1).coordinate}:{sheet.cell(row=sheet.max_row, column=col_idx-1).coordinate}"
            sheet.cell(row=2, column=col_idx).value = f"=SUM({value_range})"

            # Adding the SUMIFS formula for summing values from the last 7 days
            # It checks if the date in the date column is within the past 7 days
            date_range = f"{sheet.cell(row=2, column=date_col_idx).coordinate}:{sheet.cell(row=sheet.max_row, column=date_col_idx).coordinate}"
            sum_previous_week_formula = (
                f'=SUMIFS({value_range}, {date_range}, ">="&TODAY()-7, {date_range}, "<="&TODAY())'
            )

            # Place the sum of the previous week two cells below the total sum (i.e., in row 4)
            sheet.cell(row=4, column=col_idx).value = sum_previous_week_formula

# Save the workbook
workbook.save(filename=file_path)

# Load the dataframe to get column headers
df = pd.read_excel(file_path)

# Get a list of columns that do not have headers containing "Date", "Total", or "Other"
columns = [col for col in df.columns if "Date" not in col and "Total" not in col and "Other" not in col]

# Print the list of columns
print('\n'.join(columns))